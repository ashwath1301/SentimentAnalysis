# app.py
import streamlit as st # Needs to be imported early
import pandas as pd
import numpy as np
import os
import re
import time
import matplotlib.pyplot as plt
import seaborn as sns
from wordcloud import WordCloud, STOPWORDS
from collections import Counter
import string
from matplotlib.ticker import MaxNLocator
import nltk
from nltk.sentiment import SentimentIntensityAnalyzer
from nltk.corpus import stopwords
from nltk.tokenize import word_tokenize
from nltk.util import ngrams
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from datetime import datetime
import io # To handle byte streams from uploaded files
import traceback # For detailed error logging

# --- PAGE CONFIG (MUST BE THE FIRST STREAMLIT COMMAND) ---
st.set_page_config(layout="wide", page_title="Emerson Survey Analysis") # Changed title

# --- NLTK Setup & Global Initializations ---
@st.cache_resource
def setup_nltk():
    try: 
        nltk.data.find('sentiment/vader_lexicon.zip')
    except LookupError:
        try: 
            nltk.download('vader_lexicon', quiet=True)
        except Exception as e: 
            st.error(f"Failed to download vader_lexicon: {e}. Cannot proceed without it.")
            st.stop()
    
    try: 
        nltk.data.find('corpora/stopwords')
    except LookupError:
        try: 
            nltk.download('stopwords', quiet=True)
        except Exception as e: 
            st.warning(f"Failed to download stopwords: {e}.")
    
    try: 
        nltk.data.find('tokenizers/punkt')
    except LookupError:
        try: 
            nltk.download('punkt', quiet=True)
        except Exception as e: 
            st.error(f"Failed to download punkt tokenizer: {e}. Cannot proceed.")
            st.stop()

    # --- Updated Block for 'punkt_tab' with better error handling ---
    try: 
        nltk.data.find('tokenizers/punkt_tab')
    except (LookupError, OSError, FileNotFoundError):
        # If punkt_tab doesn't exist, try to download it
        try: 
            nltk.download('punkt_tab', quiet=True)
        except Exception as e:
            # If punkt_tab fails, try the regular punkt tokenizer
            st.warning(f"Could not download punkt_tab: {e}. Using regular punkt tokenizer.")
            try:
                nltk.data.find('tokenizers/punkt')
            except (LookupError, OSError, FileNotFoundError):
                st.error("Neither punkt_tab nor punkt tokenizer available. Tokenization might fail.")
    # --- END BLOCK ---

    try:
        sia_global = SentimentIntensityAnalyzer()
        try: 
            stop_words_global = set(stopwords.words('english'))
        except LookupError: 
            st.warning("Stopwords corpus missing.")
            stop_words_global = set()
        return sia_global, stop_words_global
    except Exception as init_e: 
        st.error(f"Failed to initialize NLTK: {init_e}")
        st.stop()

sia, stop_words_nltk = setup_nltk()

# --- File Processing Functions ---
def read_excel_from_upload(uploaded_file):
    try: uploaded_file.seek(0); return pd.read_excel(uploaded_file, engine='openpyxl')
    except Exception as e: st.error(f"Read Error {uploaded_file.name}: {e}"); return None

def read_combine_quarterly_files(uploaded_files_list, base_key_name):
    dfs = []; quarters_found = set()
    if not uploaded_files_list: return None
    for f in uploaded_files_list:
         df_q = read_excel_from_upload(f)
         if df_q is not None:
              q_match = re.search(r'[_-]?Q(\d)', f.name, re.IGNORECASE); quarter_str = f"Q{q_match.group(1)}" if q_match else "Unknown"
              if quarter_str != "Unknown": quarters_found.add(quarter_str)
              df_q['Quarter'] = quarter_str; dfs.append(df_q)
    if len(dfs) == 4 and len(quarters_found) == 4: return pd.concat(dfs, ignore_index=True)
    else: st.warning(f"{base_key_name}: Expected 4 distinct quarters, found {len(quarters_found)} ({len(dfs)} files read). Cannot combine."); return None

def save_excel_to_buffer(df, engine='openpyxl'):
    buffer = io.BytesIO()
    try:
        with pd.ExcelWriter(buffer, engine=engine) as writer: df.to_excel(writer, index=False, sheet_name='Sheet1')
        buffer.seek(0); return buffer
    except Exception as e: st.error(f"Buffer Save Error: {e}"); return None

def adjust_excel_column_widths(excel_buffer, max_width=50):
    if not excel_buffer: return None
    try:
        excel_buffer.seek(0); wb = load_workbook(excel_buffer); ws = wb.active
        for col_idx in range(1, ws.max_column + 1):
            col_letter = get_column_letter(col_idx); max_length = 0
            for row_idx in range(1, ws.max_row + 1):
                 cell = ws.cell(row=row_idx, column=col_idx);
                 try:
                     if cell.value is not None: max_length = max(max_length, len(str(cell.value)))
                 except Exception: pass
            ws.column_dimensions[col_letter].width = min(max_length + 2, max_width)
        output_buffer = io.BytesIO(); wb.save(output_buffer); output_buffer.seek(0); return output_buffer
    except Exception as e: st.warning(f"Col Width Adjust Error: {e}"); excel_buffer.seek(0); return excel_buffer

# -----------------------------------------------------------------------------
# Cleaning Functions (a - h) - Kept implementations as before
# -----------------------------------------------------------------------------
# Cleaning Function (a) - FAT
def clean_fat_data(df):
    if df is None: return None
    try:
        df = df.copy(); df.insert(0, "unique_id", range(1, len(df) + 1))
        if "Emerson Office" in df.columns: df["Emerson Office"] = df["Emerson Office"].fillna("Unknown Office")
        if "FAT Type" in df.columns:
            df["FAT Type"] = df["FAT Type"].fillna("Unknown").astype(str)
            df["FAT Type"] = df["FAT Type"].str.replace(r'\s*\(\s*Hardware\s*&\s*Software\s*\)3$', '', regex=True).str.strip()
        redacted_cols = {"Project Name": "Redacted","Customer Info": "Redacted","PM": "Redacted","Emerson Engineers": "Redacted","Internal Notes": "No internal notes provided"}
        for col, placeholder in redacted_cols.items():
            if col in df.columns and df[col].isnull().all(): df[col] = placeholder
        if "Milestones" in df.columns and "Compare" in df.columns:
             try:
                 milestones_idx, compare_idx = df.columns.get_loc("Milestones"), df.columns.get_loc("Compare")
                 if milestones_idx <= compare_idx:
                     score_columns = df.columns[milestones_idx : compare_idx + 1]
                     sentiment_map = {"Very Dissatisfied": 1,"Dissatisfied": 2,"Somewhat Dissatisfied": 2,"Somewhat Satisfied": 3,"Satisfied": 4,"Very Satisfied": 5,"Not Applicable": pd.NA,"N/A": pd.NA,0: pd.NA,'0': pd.NA,'nan': pd.NA,'': pd.NA}
                     for col in score_columns:
                         df[col] = df[col].replace(sentiment_map); df[col] = pd.to_numeric(df[col], errors="coerce")
                         if pd.api.types.is_numeric_dtype(df[col]):
                              missing_ratio = df[col].isnull().mean()
                              if missing_ratio < 0.9 and df[col].notna().sum() > 0:
                                  median_val = df[col].median(); df[col] = df[col].fillna(median_val).round().astype("Int64")
                              elif missing_ratio >= 0.9:
                                  if df[col].notna().any() and (df[col].dropna() % 1 == 0).all(): df[col] = df[col].astype("Int64")
             except KeyError: pass
             except Exception: pass
        comment_cols = ["Customer Comment", "Internal Notes"]
        for col in comment_cols:
            if col in df.columns: df[col] = df[col].fillna("No comment provided").astype(str)
        if "Submit Date" in df.columns:
            df["Submit Date"] = pd.to_datetime(df["Submit Date"], errors="coerce")
            if pd.api.types.is_datetime64_any_dtype(df["Submit Date"]):
                df["Submit Month"] = df["Submit Date"].dt.month.astype('Int64'); df["Submit Year"] = df["Submit Date"].dt.year.astype('Int64'); df["Submit Quarter"] = df["Submit Date"].dt.quarter.astype('Int64')
            else: df["Submit Month"], df["Submit Year"], df["Submit Quarter"] = pd.NA, pd.NA, pd.NA
        return df
    except Exception as e: st.error(f"Error cleaning FAT data: {e}"); return None

# Cleaning Function (b) - AEIC
def clean_aeic_data(df):
    if df is None: return None
    try:
        df = df.copy(); df.dropna(how="all", inplace=True)
        if "Completion time" not in df.columns: df["Submit Quarter"], df["Submit Quarter_Num"] = "Unknown", pd.NA
        else:
             df["Completion time"] = pd.to_datetime(df["Completion time"], errors='coerce')
             response_cutoff = df[df["Completion time"].notna()].index.max()
             if not pd.isna(response_cutoff): df = df.iloc[:response_cutoff + 1].copy()
             if pd.api.types.is_datetime64_any_dtype(df["Completion time"]):
                  df["Submit Quarter"] = df["Completion time"].dt.to_period("Q").astype(str); df["Submit Quarter_Num"] = df["Completion time"].dt.quarter.astype('Int64')
             else: df["Submit Quarter"], df["Submit Quarter_Num"] = "Unknown", pd.NA
        df.insert(0, "unique_id", range(1, len(df) + 1))
        if "Please enter your name" in df.columns: df = df.drop(columns=["Please enter your name"])
        if len(df.columns) >= 8:
             redacted_cols_names = df.columns[4:8];
             for col in redacted_cols_names: df[col] = df[col].fillna("Redacted")
        numerical_cols_indices = [12] + list(range(14, 21)) + list(range(23, 35)); numerical_cols_names = []
        for idx in numerical_cols_indices:
            if idx < len(df.columns): numerical_cols_names.append(df.columns[idx])
        for col in numerical_cols_names:
            try:
                df[col] = pd.to_numeric(df[col], errors='coerce')
                if pd.api.types.is_numeric_dtype(df[col]) and df[col].notna().sum() > 0:
                     median_val = df[col].median(); df[col] = df[col].fillna(median_val)
                     if df[col].notna().all() and (df[col] % 1 == 0).all(): df[col] = df[col].astype("Int64")
            except Exception: pass
        text_cols = df.select_dtypes(include=["object"]).columns.tolist()
        numeric_col_set = set(numerical_cols_names); text_cols_to_fill = [col for col in text_cols if col not in numeric_col_set]
        for col in text_cols_to_fill: df[col] = df[col].fillna("No comment provided")
        target_index_to_drop = 31
        if len(df) > target_index_to_drop:
            row_to_check = df.iloc[target_index_to_drop]
            if row_to_check.isnull().all() or row_to_check.astype(str).str.contains("Total|average|sum", case=False, na=False).any():
                df = df.drop(index=target_index_to_drop).reset_index(drop=True)
        return df
    except Exception as e: st.error(f"Error cleaning AEIC data: {e}"); return None

# Cleaning Function (c) - CSA Systems
def clean_csa_systems_data(df):
    if df is None: return None
    try:
        df = df.copy()
        if 'Quarter' in df.columns: df.rename(columns={"Quarter": "Submit Quarter"}, inplace=True)
        elif 'Submit Quarter' not in df.columns: df['Submit Quarter'] = 'Unknown'
        df.dropna(axis=1, how="all", inplace=True); df.insert(0, "unique_id", range(1, len(df) + 1))
        text_cols = df.select_dtypes(include=["object"]).columns.tolist()
        for col in text_cols:
            if col != 'Submit Quarter': df[col] = df[col].fillna("No comment provided")
        for col in df.columns:
            if col in ["unique_id", "Submit Quarter"]: continue
            should_try_numeric = False; col_dtype = df[col].dtype
            if pd.api.types.is_numeric_dtype(col_dtype): should_try_numeric = True
            elif col_dtype == 'object':
                 non_null_sample = df[col].dropna()
                 if not non_null_sample.empty and non_null_sample.astype(str).str.match(r'^-?\d+(\.\d+)?$').all(): should_try_numeric = True
            if should_try_numeric:
                 try:
                     df[col] = pd.to_numeric(df[col], errors="coerce")
                     if pd.api.types.is_numeric_dtype(df[col]) and df[col].notna().sum() > 0:
                         median_val = df[col].median(); df[col] = df[col].fillna(median_val)
                         if df[col].notna().all() and (df[col] % 1 == 0).all(): df[col] = df[col].astype("Int64")
                 except Exception: pass
        if "Submit Quarter" in df.columns:
             df["Submit_Quarter_Num"] = df["Submit Quarter"].astype(str).str.extract(r'Q(\d)', expand=False)
             df["Submit_Quarter_Num"] = pd.to_numeric(df["Submit_Quarter_Num"], errors='coerce').astype('Int64'); df["Half"] = pd.NA
             df.loc[df["Submit_Quarter_Num"].isin([1, 2]), "Half"] = "H1"; df.loc[df["Submit_Quarter_Num"].isin([3, 4]), "Half"] = "H2"
        return df
    except Exception as e: st.error(f"Error cleaning CSA Systems data: {e}"); return None

# Cleaning Function (d) - Ovation
def clean_ovation_product_support_data(df):
    if df is None: return None
    try:
        df = df.copy()
        if 'Quarter' in df.columns: df.rename(columns={"Quarter": "Submit Quarter"}, inplace=True)
        elif 'Submit Quarter' not in df.columns: df['Submit Quarter'] = 'Unknown'
        if "Quarter" in df.columns and df['Quarter'].dtype != 'object': df = df.drop(columns=["Quarter"])
        df.dropna(axis=1, how="all", inplace=True); df.insert(0, "unique_id", range(1, len(df) + 1))
        text_cols = df.select_dtypes(include=["object"]).columns.tolist()
        for col in text_cols:
             if col != 'Submit Quarter': df[col] = df[col].fillna("No comment provided")
        for col in df.columns:
            if col in ["unique_id", "Submit Quarter"]: continue
            should_try_numeric = False; col_dtype = df[col].dtype
            if pd.api.types.is_numeric_dtype(col_dtype): should_try_numeric = True
            elif col_dtype == 'object':
                 non_null_sample = df[col].dropna();
                 if not non_null_sample.empty and non_null_sample.astype(str).str.match(r'^-?\d+(\.\d+)?$').all(): should_try_numeric = True
            if should_try_numeric:
                 try:
                     df[col] = pd.to_numeric(df[col], errors="coerce")
                     if pd.api.types.is_numeric_dtype(df[col]) and df[col].notna().sum() > 0:
                         median_val = df[col].median(); df[col] = df[col].fillna(median_val)
                         if df[col].notna().all() and (df[col] % 1 == 0).all(): df[col] = df[col].astype("Int64")
                 except Exception: pass
        return df
    except Exception as e: st.error(f"Error cleaning Ovation data: {e}"); return None

# Cleaning Function (e) - Material Replacement
def clean_material_replacement_data(df):
    if df is None: return None
    try:
         df = df.copy()
         if 'Quarter' in df.columns: df.rename(columns={"Quarter": "Submit Quarter"}, inplace=True)
         elif 'Submit Quarter' not in df.columns: df['Submit Quarter'] = 'Unknown';
         if "Quarter" in df.columns and df['Quarter'].dtype != 'object': df = df.drop(columns=["Quarter"])
         df.dropna(axis=1, how="all", inplace=True); df.insert(0, "unique_id", range(1, len(df) + 1))
         text_cols = df.select_dtypes(include=["object"]).columns.tolist()
         for col in text_cols:
              if col != 'Submit Quarter': df[col] = df[col].fillna("No comment provided")
         for col in df.columns:
             if col in ["unique_id", "Submit Quarter"]: continue
             should_try_numeric = False; col_dtype = df[col].dtype
             if pd.api.types.is_numeric_dtype(col_dtype): should_try_numeric = True
             elif col_dtype == 'object':
                  non_null_sample = df[col].dropna();
                  if not non_null_sample.empty and non_null_sample.astype(str).str.match(r'^-?\d+(\.\d+)?$').all(): should_try_numeric = True
             if should_try_numeric:
                  try:
                      df[col] = pd.to_numeric(df[col], errors="coerce")
                      if pd.api.types.is_numeric_dtype(df[col]) and df[col].notna().sum() > 0:
                          median_val = df[col].median(); df[col] = df[col].fillna(median_val)
                          if df[col].notna().all() and (df[col] % 1 == 0).all(): df[col] = df[col].astype("Int64")
                  except Exception: pass
         return df
    except Exception as e: st.error(f"Error cleaning Material Repl data: {e}"); return None

# Cleaning Function (f) - Field Service
def clean_field_service_data(df):
    if df is None: return None
    try:
        df = df.copy()
        if 'Quarter' in df.columns: df.rename(columns={"Quarter": "Submit Quarter"}, inplace=True)
        elif 'Submit Quarter' not in df.columns: df['Submit Quarter'] = 'Unknown'
        if "Quarter" in df.columns and df['Quarter'].dtype != 'object': df = df.drop(columns=["Quarter"])
        df.dropna(axis=1, how="all", inplace=True); df.insert(0, "unique_id", range(1, len(df) + 1))
        text_cols = df.select_dtypes(include=["object"]).columns.tolist()
        for col in text_cols:
             if col != 'Submit Quarter': df[col] = df[col].fillna("No comment provided")
        for col in df.columns:
            if col in ["unique_id", "Submit Quarter"]: continue
            should_try_numeric = False; col_dtype = df[col].dtype
            if pd.api.types.is_numeric_dtype(col_dtype): should_try_numeric = True
            elif col_dtype == 'object':
                 non_null_sample = df[col].dropna();
                 if not non_null_sample.empty and non_null_sample.astype(str).str.match(r'^-?\d+(\.\d+)?$').all(): should_try_numeric = True
            if should_try_numeric:
                 try:
                     df[col] = pd.to_numeric(df[col], errors="coerce")
                     if pd.api.types.is_numeric_dtype(df[col]) and df[col].notna().sum() > 0:
                         median_val = df[col].median(); df[col] = df[col].fillna(median_val)
                         if df[col].notna().all() and (df[col] % 1 == 0).all(): df[col] = df[col].astype("Int64")
                 except Exception: pass
        return df
    except Exception as e: st.error(f"Error cleaning Field Service data: {e}"); return None

# Cleaning Function (g) - Installed Base
def clean_installed_base_data(df):
    if df is None: return None
    try:
        df = df.copy(); df.insert(0, "unique_id", range(1, len(df) + 1))
        if "Call Date" in df.columns:
             df["Call Date"] = pd.to_datetime(df["Call Date"], format="%Y%m%d", errors='coerce')
             if not pd.api.types.is_datetime64_any_dtype(df["Call Date"]) or df["Call Date"].isnull().all(): df["Call Date"] = pd.to_datetime(df["Call Date"], errors='coerce')
             if pd.api.types.is_datetime64_any_dtype(df["Call Date"]) and df['Call Date'].notna().any():
                  df["Submit_Quarter_Num"] = df["Call Date"].dt.quarter.astype('Int64'); df["Submit Quarter"] = df["Call Date"].dt.year.astype(str).replace('nan', 'Unknown') + "Q" + df["Submit_Quarter_Num"].astype(str).replace('<NA>', '?'); df["Submit Quarter"] = df["Submit Quarter"].replace({'UnknownQ?': 'Unknown'})
             else: df["Submit_Quarter_Num"], df["Submit Quarter"] = pd.NA, "Unknown"
        else: df["Submit_Quarter_Num"], df["Submit Quarter"] = pd.NA, "Unknown"
        if "Quarter" in df.columns and df['Quarter'].dtype != 'object': df = df.drop(columns=["Quarter"])
        if "Year" in df.columns:
             original_type = df['Year'].dtype; df["Year"] = pd.to_numeric(df["Year"], errors='coerce'); df["Year"] = df["Year"].replace(7, 2024)
             if df["Year"].notna().all() and (df["Year"] % 1 == 0).all(): df["Year"] = df["Year"].astype('Int64')
             elif not pd.api.types.is_numeric_dtype(df['Year']): df["Year"] = df["Year"].astype(original_type)
        comment_cols_letters = ["Q","AC","AF","AN","AR","AT","AU","AZ","BC","BE","BG","BJ","BL","BN","BQ","BP"]
        comment_cols_present = [col for col in comment_cols_letters if col in df.columns]
        for col in comment_cols_present: df[col] = df[col].fillna("No comment provided").astype(str).replace('nan', 'No comment provided')
        return df
    except Exception as e: st.error(f"Error cleaning Inst Base data: {e}"); return None

# Cleaning Function (h) - NA Users Written
def clean_na_users_written_data(uploaded_file):
    if uploaded_file is None: return None
    try:
        uploaded_file.seek(0); xls = pd.ExcelFile(uploaded_file); all_comments = []; processed_sheets_count = 0
        for sheet_name in xls.sheet_names:
            try:
                df_sheet = pd.read_excel(xls, sheet_name=sheet_name); df_sheet.dropna(how="all", inplace=True)
                if len(df_sheet.columns) >= 2 and not df_sheet.empty:
                     id_col, comment_col = df_sheet.columns[0], df_sheet.columns[1]; question_text = str(comment_col)
                     if (pd.api.types.is_numeric_dtype(df_sheet[id_col].dropna()) or df_sheet[id_col].dtype == 'object') and df_sheet[comment_col].dtype == 'object':
                         temp_df = df_sheet[[id_col, comment_col]].copy(); temp_df.columns = ["unique_id", "comment"]; temp_df["question"] = question_text; all_comments.append(temp_df); processed_sheets_count += 1
            except Exception: pass
        if not all_comments: return None
        combined_df = pd.concat(all_comments, ignore_index=True)
        combined_df["comment"] = combined_df["comment"].fillna("No comment provided").astype(str); combined_df['unique_id'] = combined_df['unique_id'].astype(str)
        return combined_df
    except Exception as e: st.error(f"Error cleaning NA Users Written data: {e}"); return None

# -----------------------------------------------------------------------------
# Sentiment Analysis Function
# -----------------------------------------------------------------------------
def apply_sentiment_analysis(df):
    if df is None: return None
    try:
        df = df.copy(); fallback_patterns = ["comment","notes","suggestion","feedback","additional","reason","O_S5A","O_GEN_COM","O_Q7","O_Q9a","O_Q10a","O_Q18","O_Q20a","O_Q25","O_Q27","O_Q29a","O_Q31a","O_Q32a","O_Q33a","O_Q34a","O_Q35","Q5","Q6","Q11","Q14a","Q15a","Q16a","Q18a","Q18b","Q19a","Q20a","Q2a","Q4a","Q11a","Q13","Q15","Q17","Q18","Q22a","Q24a","Q25a","Q26a","Q28a","Q29a","S4a","S5a","Q7","Q8","Q10a","Q12a","Q13a","Q14a","Q17a","Q9","Q10","Q12a","Q14a","Q15a","Q16a","Q19a","Q","AC","AF","AN","AR","AT","AU","AZ","BC","BE","BG","BJ","BL","BN","BQ","BP"]
        patterns_lower = [p.lower() for p in fallback_patterns]; exact_matches = ["comment"]; comment_cols = []
        for col in df.columns:
            col_str = str(col); col_lower = col_str.lower()
            if df[col].dtype == 'object':
                 is_match = False
                 if col_lower in exact_matches: is_match = True
                 else:
                      for pattern in patterns_lower:
                           if pattern in col_lower: is_match = True; break
                 if is_match:
                      non_null_sample = df[col].dropna()
                      if non_null_sample.empty or non_null_sample.astype(str).str.contains(' ').any(): comment_cols.append(col_str)
        comment_cols = sorted(list(set(comment_cols)));
        if not comment_cols: return df
        sentiment_applied_count = 0
        for col in comment_cols:
             sentiment_score_col, label_col = f"{col}_SentimentScore", f"{col}_SentimentLabel"
             if label_col in df.columns: continue
             try:
                 df[col] = df[col].fillna("").astype(str)
                 sentiment_scores = df[col].apply(lambda x: sia.polarity_scores(x)["compound"] if isinstance(x, str) else 0)
                 df[sentiment_score_col] = sentiment_scores
                 df[label_col] = sentiment_scores.apply(lambda s: "Positive" if s > 0.05 else ("Negative" if s < -0.05 else "Neutral"))
                 sentiment_applied_count += 1
             except Exception: pass
        return df
    except Exception as e: st.error(f"Error applying sentiment analysis: {e}"); return df

# -----------------------------------------------------------------------------
# Sentiment Summarization Function
# -----------------------------------------------------------------------------
def generate_sentiment_summary(df, groupby_col="Submit Quarter"):
    if df is None: return None
    try:
        df = df.copy(); sentiment_label_cols = [col for col in df.columns if str(col).lower().endswith("_sentimentlabel")]
        if not sentiment_label_cols: return pd.DataFrame()
        summary_frames = []; processed_base_metrics = set()
        potential_group_cols = [groupby_col, 'Submit_Quarter_Num', 'Quarter', 'Submit Quarter_Num']
        actual_groupby_col = None
        for potential_col in potential_group_cols:
            if potential_col in df.columns and df[potential_col].nunique() > 1: actual_groupby_col = potential_col; break
        for col in sentiment_label_cols:
            base_metric = re.sub(r'_sentimentlabel$', '', col, flags=re.IGNORECASE); base_metric = re.sub(r'^[OoSs]_', '', base_metric); base_metric = base_metric.replace('_', ' ').strip()
            if base_metric.lower() in processed_base_metrics: continue
            processed_base_metrics.add(base_metric.lower())
            if actual_groupby_col:
                try:
                    grouped = df.groupby(actual_groupby_col)[col].value_counts(normalize=True).unstack(fill_value=0) * 100; grouped = grouped.round(2)
                    for sentiment in ["Positive", "Neutral", "Negative"]:
                        if sentiment not in grouped.columns: grouped[sentiment] = 0.0
                    grouped = grouped[["Positive", "Neutral", "Negative"]]; grouped.columns = [f"{base_metric} - {label} %" for label in grouped.columns]; grouped.reset_index(inplace=True)
                    summary_frames.append(grouped)
                except Exception: pass
            else:
                try:
                    overall_dist = df[col].value_counts(normalize=True) * 100; overall_dist = overall_dist.round(2)
                    for sentiment in ["Positive", "Neutral", "Negative"]:
                         if sentiment not in overall_dist.index: overall_dist[sentiment] = 0.0
                    overall_dist = overall_dist.reindex(["Positive", "Neutral", "Negative"]).fillna(0.0)
                    summary_df = pd.DataFrame({"Metric": [base_metric] * len(overall_dist), "Sentiment": overall_dist.index, "Percentage": overall_dist.values})
                    summary_frames.append(summary_df)
                except Exception: pass
        if not summary_frames: return pd.DataFrame()
        if actual_groupby_col:
            if len(summary_frames) == 0: return pd.DataFrame()
            if len(summary_frames) == 1: return summary_frames[0]
            final_summary = summary_frames[0]
            for i in range(1, len(summary_frames)):
                 if actual_groupby_col in summary_frames[i].columns: final_summary = pd.merge(final_summary, summary_frames[i], on=actual_groupby_col, how="outer")
            for col in final_summary.columns:
                 if col != actual_groupby_col and pd.api.types.is_numeric_dtype(final_summary[col]): final_summary[col] = final_summary[col].fillna(0)
            return final_summary
        else: return pd.concat(summary_frames, ignore_index=True)
    except Exception as e: st.error(f"Error generating sentiment summary: {e}"); return pd.DataFrame()

# --- Visualization Functions ---

# FAT Survey Visualizations (Keep as in previous version)
def display_fat_visualizations(summary_file_buffer, sentiment_file_buffer):
    st.header("FAT Survey Analysis")
    try:
        summary_file_buffer.seek(0); sentiment_file_buffer.seek(0)
        df_area = pd.read_excel(summary_file_buffer); df_sentiment = pd.read_excel(sentiment_file_buffer)
        df_area = df_area.copy(); df_sentiment = df_sentiment.copy()
        st.subheader("Customer Feedback Trends Over Time (Quarterly Sentiment)")
        df_area.columns = df_area.columns.str.strip(); qtr_col = None; qtr_col_options = ['Submit Quarter', 'Submit_Quarter_Num', 'Submit Quarter_Num']
        for col in qtr_col_options:
             if col in df_area.columns and pd.api.types.is_numeric_dtype(df_area[col].dropna()):
                  if (df_area[col].dropna() % 1 == 0).all(): df_area[col] = df_area[col].astype('Int64'); qtr_col = col; break
        if qtr_col is None and 'Submit Quarter' in df_area.columns and df_area['Submit Quarter'].dtype == 'object':
              df_area['QuarterNum_extracted'] = df_area['Submit Quarter'].astype(str).str.extract(r'Q(\d)', expand=False); df_area['QuarterNum_extracted'] = pd.to_numeric(df_area['QuarterNum_extracted'], errors='coerce').astype('Int64')
              if df_area['QuarterNum_extracted'].notna().any(): qtr_col = 'QuarterNum_extracted'
        if qtr_col is None: st.error("FAT Viz: Failed to find quarter column."); return
        sent_cols_map = {}; possible_bases = ["Customer Comment", "Comment", "Feedback"]; sentiments = ["Positive", "Neutral", "Negative"]
        for base in possible_bases:
            for sent in sentiments:
                 pattern = f"{base}.*{sent}.*%";
                 for col in df_area.columns:
                     if re.search(pattern, col, re.IGNORECASE) and sent not in sent_cols_map: sent_cols_map[sent] = col
        for sent in sentiments:
            if sent not in sent_cols_map:
                 pattern = f"{sent}.*%"
                 for col in df_area.columns:
                      if re.search(pattern, col, re.IGNORECASE) and sent not in sent_cols_map: sent_cols_map[sent] = col
        if not all(s in sent_cols_map for s in sentiments): st.error(f"FAT Viz: Missing sentiment % columns."); return
        df_area.rename(columns={v: k for k, v in sent_cols_map.items()}, inplace=True)
        df_area['x_val'] = df_area[qtr_col].map({1: 0, 2: 1, 3: 2, 4: 3}); df_area_plot = df_area.dropna(subset=['x_val', 'Positive', 'Neutral', 'Negative']).copy()
        df_area_plot['x_val'] = df_area_plot['x_val'].astype(int)
        if df_area_plot.empty: st.warning("FAT Viz: No valid quarterly data for Area Chart."); return
        df_area_plot = df_area_plot.groupby('x_val')[['Positive', 'Neutral', 'Negative']].mean().reindex(range(4), fill_value=0)
        fig_area, ax_area = plt.subplots(figsize=(10, 6))
        df_area_plot.plot.area(ax=ax_area, stacked=True, alpha=0.85, color=["#4CAF50", "#FFB300", "#F44336"])
        label_threshold = 2.0; current_y = pd.Series(0.0, index=df_area_plot.index)
        for sentiment in ["Positive", "Neutral", "Negative"]:
            for x, val in df_area_plot[sentiment].items():
                 if val >= label_threshold: ax_area.text(x, current_y[x] + val / 2, f'{val:.1f}%', color='black', ha='center', va='center', fontsize=9, weight='bold')
            current_y += df_area_plot[sentiment]
        ax_area.set_xlim(-0.5, 3.5); ax_area.set_ylim(0, 100); ax_area.set_xticks([0, 1, 2, 3]); ax_area.set_xticklabels(["Q1", "Q2", "Q3", "Q4"])
        ax_area.set_title("FAT: Customer Feedback Trends Over Time", fontsize=16, fontweight='bold'); ax_area.set_xlabel("Timeline (Quarterly)", fontsize=12); ax_area.set_ylabel("Sentiment Composition (%)", fontsize=12)
        ax_area.legend(title="Sentiment", loc="upper left", bbox_to_anchor=(1.01, 1)); ax_area.grid(axis='y', alpha=0.3)
        plt.tight_layout(rect=[0, 0, 0.85, 1]); st.pyplot(fig_area); plt.close(fig_area)
        st.subheader("Sentiment Breakdown by Category"); df_sentiment.columns = df_sentiment.columns.str.strip()
        sentiment_label_cols = [col for col in df_sentiment.columns if str(col).lower().endswith("_sentimentlabel")]; main_sentiment_col = None
        if sentiment_label_cols: cust_comment_col = [c for c in sentiment_label_cols if "customer comment" in c.lower()]; main_sentiment_col = cust_comment_col[0] if cust_comment_col else sentiment_label_cols[0]
        if not main_sentiment_col: st.warning("FAT Viz: No Sentiment Label for breakdown."); return
        category_cols = ["FAT Type", "Emerson Office"]; category_cols_present = [col for col in category_cols if col in df_sentiment.columns]
        if not category_cols_present: st.warning("FAT Viz: No category columns for breakdown."); return
        melted = df_sentiment.melt(id_vars=category_cols_present, value_vars=[main_sentiment_col], var_name="Question", value_name="Sentiment"); melted.dropna(subset=['Sentiment'], inplace=True)
        custom_palette_type = {"Positive": "#3CB371", "Neutral": "#F0E68C", "Negative": "#CD5C5C"}; custom_palette_office = {"Positive": "#228B22", "Neutral": "#FFD700", "Negative": "#DC143C"}
        if "FAT Type" in category_cols_present:
             st.write("#### Sentiment by Testing Type")
             if df_sentiment["FAT Type"].isnull().all() or df_sentiment["FAT Type"].nunique() == 0: st.info("No data in 'FAT Type'.")
             else:
                  fat_grouped = melted.groupby(["FAT Type", "Sentiment"]).size().unstack(fill_value=0); fat_perc = fat_grouped.apply(lambda x: x / x.sum() * 100 if x.sum() > 0 else x, axis=1)
                  for sent in ["Positive", "Neutral", "Negative"]:
                      if sent not in fat_perc.columns: fat_perc[sent] = 0
                  fat_perc = fat_perc[["Positive", "Neutral", "Negative"]]; fig_type, ax_type = plt.subplots(figsize=(10, 6))
                  colors_type = [custom_palette_type.get(c, '#808080') for c in fat_perc.columns]; fat_perc.plot(kind="bar", stacked=True, color=colors_type, ax=ax_type, width=0.8)
                  ax_type.set_title("FAT: Sentiment Distribution by Testing Type", fontsize=14, fontweight='bold'); ax_type.set_ylabel("Percentage (%)"); ax_type.set_xlabel("FAT Type")
                  ax_type.tick_params(axis='x', labelsize=9); plt.setp(ax_type.get_xticklabels(), rotation=45, ha='right', rotation_mode='anchor')
                  ax_type.set_ylim(0, 100); ax_type.legend(title="Sentiment", loc="upper left", bbox_to_anchor=(1.01, 1.0))
                  plt.tight_layout(rect=[0, 0.1, 0.85, 1]); st.pyplot(fig_type); plt.close(fig_type)
        if "Emerson Office" in category_cols_present:
            st.write("#### Sentiment by Distribution Center")
            if df_sentiment["Emerson Office"].isnull().all() or df_sentiment["Emerson Office"].nunique() == 0: st.info("No data in 'Emerson Office'.")
            else:
                office_grouped = melted.groupby(["Emerson Office", "Sentiment"]).size().unstack(fill_value=0); office_perc = office_grouped.apply(lambda x: x / x.sum() * 100 if x.sum() > 0 else x, axis=1)
                for sent in ["Positive", "Neutral", "Negative"]:
                    if sent not in office_perc.columns: office_perc[sent] = 0
                office_perc = office_perc[["Positive", "Neutral", "Negative"]]; fig_office, ax_office = plt.subplots(figsize=(max(8, len(office_perc)*0.8), 6))
                colors_office = [custom_palette_office.get(c, '#808080') for c in office_perc.columns]; office_perc.plot(kind="bar", stacked=True, color=colors_office, ax=ax_office, width=0.8)
                ax_office.set_title("FAT: Sentiment Distribution by Distribution Center", fontsize=14, fontweight='bold'); ax_office.set_ylabel("Percentage (%)"); ax_office.set_xlabel("Distribution Center")
                ax_office.tick_params(axis='x', labelsize=max(6, 10 - len(office_perc)//5)); plt.setp(ax_office.get_xticklabels(), rotation=45, ha='right', rotation_mode='anchor')
                ax_office.set_ylim(0, 100); ax_office.legend(title="Sentiment", loc="upper left", bbox_to_anchor=(1.01, 1.0))
                plt.tight_layout(rect=[0, 0.1, 0.85, 1]); st.pyplot(fig_office); plt.close(fig_office)
    except Exception as e: st.error(f"Unexpected error during FAT visualization: {e}"); st.error(traceback.format_exc())

# AEIC Survey Visualizations (Chart 1 reverted, Chart 2 verified, Chart 3 title corrected)
def display_aeic_visualizations(sentiment_file_buffer):
    st.header("AEIC Survey Analysis")
    try:
        sentiment_file_buffer.seek(0); df = pd.read_excel(sentiment_file_buffer); df = df.copy()
        df.columns = df.columns.str.replace("\xa0", " ", regex=False).str.strip()
        st.subheader("Sentiment Distribution & Ratings")

        # --- CHART 1: Sentiment Distribution Bar Chart (REVERTED TO USER'S ORIGINAL CODE) ---
        st.write("#### Sentiment Distribution by Comment Theme") # Add sub-subheader
        sentiment_cols = [col for col in df.columns if col.endswith("_SentimentLabel")]
        if len(sentiment_cols) <3: st.warning(f"AEIC Viz Chart 1: Expected 3 sentiment columns, found {len(sentiment_cols)}. Using available."); rename_map = {col: re.sub(r'_sentimentlabel$', '', col, flags=re.IGNORECASE).strip() for col in sentiment_cols}
        else: rename_map = { sentiment_cols[0]: "Customer Satisfaction", sentiment_cols[1]: "Supplier Value", sentiment_cols[2]: "Customer Suggestions for Focus Areas" }
        cols_to_select = [col for col in rename_map.keys() if col in df.columns]
        if not cols_to_select: st.warning("AEIC Viz Chart 1: None of the expected sentiment columns found."); return
        df_melted = df[cols_to_select].copy(); df_melted.rename(columns=rename_map, inplace=True); mapped_cols_present = list(df_melted.columns)
        df_melted["Response ID"] = range(1, len(df_melted) + 1) # Use range for ID as per original
        df_long = df_melted.melt(id_vars="Response ID", value_vars=mapped_cols_present, var_name="Question Type", value_name="Sentiment"); df_long = df_long[df_long["Sentiment"].notna()]
        def smart_round(row): # Original smart_round function from user snippet
            if row.sum() == 0: return row.astype(int)
            raw = row / row.sum() * 100; rounded = raw.round(); diff = int(100 - rounded.sum())
            if diff != 0: remainder = (raw - rounded).abs(); target = remainder.idxmax(); rounded[target] += diff
            return rounded.astype(int)
        qtype_sentiment = df_long.groupby(["Question Type", "Sentiment"]).size().unstack().fillna(0)
        if qtype_sentiment.empty: st.warning("AEIC Viz Chart 1: No data after grouping.")
        else:
            for sent in ["Positive", "Neutral", "Negative"]:
                if sent not in qtype_sentiment.columns: qtype_sentiment[sent] = 0
            qtype_sentiment_perc = qtype_sentiment.apply(smart_round, axis=1)[["Positive", "Neutral", "Negative"]]
            labels = qtype_sentiment_perc.index.tolist(); custom_colors = ["#228B22", "#FFD700", "#DC143C"]
            fig1, ax1 = plt.subplots(figsize=(10, 6))
            qtype_sentiment_perc.plot(kind="bar", stacked=True, color=custom_colors, ax=ax1)
            for i, (idx, row) in enumerate(qtype_sentiment_perc.iterrows()):
                 pos_val, neu_val, neg_val = row["Positive"], row["Neutral"], row["Negative"]
                 if pos_val > 1: ax1.text(i, pos_val / 2, f'{pos_val:.0f}%', color="white", ha="center", va='center', fontweight="bold")
                 if neu_val > 1: ax1.text(i, pos_val + neu_val / 2, f'{neu_val:.0f}%', color="black", ha="center", va='center', fontweight="bold")
                 neg_y_pos = pos_val + neu_val + neg_val / 2 if (pos_val + neu_val) < 98 else 100 - neg_val / 2
                 if neg_val > 1: ax1.text(i, neg_y_pos, f'{neg_val:.0f}%', color="white", ha="center", va='center', fontweight="bold")
            ax1.set_xticks(range(len(labels))); ax1.set_xticklabels(labels, rotation=0, fontsize=10)
            ax1.set_title("AEIC Survey – Sentiment Distribution by Comment Theme", fontsize=16, fontweight="bold"); ax1.set_ylabel("Percentage"); ax1.set_xlabel("Comment Theme")
            ax1.legend(title="Sentiment", loc="upper left", bbox_to_anchor=(1.01, 1.0)); ax1.grid(axis="y", linestyle="--", alpha=0.4)
            plt.tight_layout(rect=[0, 0, 0.85, 1]); st.pyplot(fig1); plt.close(fig1)

        # --- CHART 2: Average Ratings Bar Chart (Verified Original Logic) ---
        st.write("#### Average Ratings") # Add sub-subheader
        numeric_cols = [col for col in df.columns if ('Supplier Value -' in col or 'Customer Satisfaction -' in col) and not str(col).lower().endswith(('_sentimentlabel', '_sentimentscore'))]
        if not numeric_cols: st.warning("AEIC Viz: No numeric rating columns found.")
        else:
            df_ratings = df[numeric_cols].apply(pd.to_numeric, errors="coerce"); avg_ratings = df_ratings.mean().dropna().sort_values()
            if avg_ratings.empty: st.warning("AEIC Viz: No valid average ratings.")
            else:
                short_labels = {col: ("Customer Satisfaction: " + col.split("Customer Satisfaction - ")[1].split(":")[0].strip() if col.startswith("Customer Satisfaction -") else "Supplier Value: " + col.split("Supplier Value - ")[1].split(":")[0].strip()) for col in avg_ratings.index}
                avg_ratings.index = [short_labels.get(col, col) for col in avg_ratings.index]
                min_val, max_val = avg_ratings.min(), avg_ratings.max(); colors = ["lightblue" if val == min_val else "red" if val == max_val else "gray" for val in avg_ratings]
                fig2, ax2 = plt.subplots(figsize=(10, max(6, len(avg_ratings)*0.4)))
                avg_ratings.plot(kind="barh", color=colors, ax=ax2)
                ax2.set_title("Average Ratings across AEIC Survey Areas", fontsize=14, fontweight="bold"); ax2.set_xlabel("Average Rating (1-10)")
                ax2.tick_params(axis='y', labelsize=9);
                for index, value in enumerate(avg_ratings): ax2.text(value + 0.05, index, f' {value:.2f}', va='center', ha='left', fontsize=8)
                ax2.grid(axis="x", linestyle="--", alpha=0.7); plt.tight_layout(); st.pyplot(fig2); plt.close(fig2)

        # --- CHART 3: Top Negative Keywords (Corrected Title) ---
        cust_sat_comment_col, cust_sat_sentiment_col = None, None
        target_col_substring = "Customer Satisfaction: To ensure maximum benefit"
        for col in df.columns:
            if target_col_substring in str(col) and not str(col).lower().endswith(("_sentimentlabel", "_sentimentscore")): cust_sat_comment_col = col; break
        if cust_sat_comment_col:
            sentiment_col_guess = cust_sat_comment_col + "_SentimentLabel";
            if sentiment_col_guess in df.columns: cust_sat_sentiment_col = sentiment_col_guess
        if not cust_sat_comment_col:
            for col in df.columns:
                col_lower = str(col).lower()
                if "customer satisfaction" in col_lower and ("comment" in col_lower or "feedback" in col_lower or "rating(s)" in col_lower) and not col_lower.endswith(("_sentimentlabel", "_sentimentscore")): cust_sat_comment_col = col
                if cust_sat_comment_col and col_lower == f"{cust_sat_comment_col.lower()}_sentimentlabel": cust_sat_sentiment_col = col; break
        if not cust_sat_comment_col or not cust_sat_sentiment_col: st.warning("AEIC Viz: Cannot find Cust Sat comment/sentiment columns.")
        else:
            keyword_chart_title = "AEIC: Top Keywords in Negative Customer Satisfaction Comments"; st.write(f"#### {keyword_chart_title}")
            custom_stopwords = stop_words_nltk | {"please","emerson","customer","satisfaction","supplier","value","comment","provided","n/a","none","na","no","comments","provide","if","get","also","would","could","need","well","us","good","like","see","think","one","make","work","support","aeic"}
            def extract_keywords_aeic(texts, top_n=15):
                word_list = [];
                for text in texts: text = str(text).lower(); text = re.sub(r'[^\w\s]', '', text); tokens = word_tokenize(text); valid_words = [word for word in tokens if word not in custom_stopwords and len(word) > 2 and not word.isdigit()]; word_list.extend(valid_words)
                return Counter(word_list).most_common(top_n)
            try: negative_comments = df.loc[df[cust_sat_sentiment_col] == "Negative", cust_sat_comment_col].dropna().astype(str).tolist()
            except KeyError: st.error(f"AEIC Viz: Error accessing keyword columns."); negative_comments = []
            if not negative_comments: st.info(f"No negative comments in '{cust_sat_comment_col}'.")
            else:
                 top_keywords = extract_keywords_aeic(negative_comments)
                 if not top_keywords: st.info("No significant keywords found.")
                 else:
                      fig3, ax3 = plt.subplots(figsize=(8, 6)); words, freqs = zip(*top_keywords); ax3.barh(words[::-1], freqs[::-1], color="#D9534F")
                      ax3.set_title(keyword_chart_title, fontsize=14, weight="bold"); ax3.set_xlabel("Frequency")
                      plt.tight_layout(); st.pyplot(fig3); plt.close(fig3)
    except Exception as e: st.error(f"Unexpected error during AEIC visualization: {e}"); st.error(traceback.format_exc())

# NA Users Group Survey Visualizations (Keep as before)
def display_na_users_visualizations(written_sentiment_buffer, numerical_template_buffer):
    st.header("NA Users Group Survey Analysis")
    try:
        written_sentiment_buffer.seek(0); numerical_template_buffer.seek(0)
        written_df = pd.read_excel(written_sentiment_buffer); numerical_df = pd.read_excel(numerical_template_buffer)
        written_df = written_df.copy(); numerical_df = numerical_df.copy()
        st.subheader("Key Areas from Written Feedback and Numerical Scores")
        fig, axs = plt.subplots(2, 1, figsize=(12, 14))
        if "comment_SentimentLabel" not in written_df.columns or "question" not in written_df.columns:
             st.warning("NA Users Viz: Missing columns in written comments."); axs[0].text(0.5, 0.5, "Data unavailable", ha='center', va='center', color='gray')
        else:
             neg_counts = written_df[written_df["comment_SentimentLabel"] == "Negative"].groupby("question", dropna=False)["comment"].count().sort_values(ascending=True); neg_counts = neg_counts[neg_counts > 0]
             if neg_counts.empty: axs[0].text(0.5, 0.5, "No negative comments", ha='center', va='center', color='gray')
             else:
                  max_label_len = 70; cleaned_labels = [(str(q)[:max_label_len-3] + "...") if len(str(q)) > max_label_len else str(q) for q in neg_counts.index]
                  axs[0].barh(cleaned_labels, neg_counts.values, color="#D9534F")
                  axs[0].xaxis.set_major_locator(MaxNLocator(integer=True, min_n_ticks=3)); axs[0].tick_params(axis='y', labelsize=9)
                  for index, value in enumerate(neg_counts.values): axs[0].text(value, index, f' {value}', va='center', ha='left', fontsize=8)
             axs[0].set_title("NA Users: Number of Negative Written Comments by Question", fontsize=14, fontweight="bold"); axs[0].set_xlabel("Count of Negative Responses"); axs[0].grid(axis='x', linestyle='--', alpha=0.6)
        req_num_cols = ["Questions", "Weighted Average Score"]
        if not all(col in numerical_df.columns for col in req_num_cols):
             st.warning(f"NA Users Viz: Missing columns ({req_num_cols}) in numerical template."); axs[1].text(0.5, 0.5, "Data unavailable", ha='center', va='center', color='gray')
        else:
             numerical_df["Weighted Average Score"] = pd.to_numeric(numerical_df["Weighted Average Score"], errors='coerce')
             lowest_scores = numerical_df[req_num_cols].dropna().sort_values("Weighted Average Score").head(10)
             if lowest_scores.empty: axs[1].text(0.5, 0.5, "No valid numerical scores", ha='center', va='center', color='gray')
             else:
                  max_label_len_num = 70; lowest_scores["Questions_Short"] = lowest_scores["Questions"].apply(lambda x: (str(x)[:max_label_len_num-3] + "...") if len(str(x)) > max_label_len_num else str(x))
                  axs[1].barh(lowest_scores["Questions_Short"], lowest_scores["Weighted Average Score"], color="#5BC0DE")
                  axs[1].tick_params(axis='y', labelsize=9)
                  for index, value in enumerate(lowest_scores["Weighted Average Score"]): axs[1].text(value, index, f' {value:.2f}', va='center', ha='left', fontsize=8)
             axs[1].set_title("NA Users: Lowest Rated Questions (Top 10 by Weighted Average Score)", fontsize=14, fontweight="bold"); axs[1].set_xlabel("Weighted Average Score"); axs[1].grid(axis='x', linestyle='--', alpha=0.6)
        plt.tight_layout(h_pad=4.0); st.pyplot(fig); plt.close(fig)
    except Exception as e: st.error(f"Unexpected error during NA Users viz: {e}"); st.error(traceback.format_exc())

# Schmidt Surveys Visualizations - USING ORIGINAL SNIPPETS (with CSA Theme fix & IB Theme remove)
def display_schmidt_visualizations_original(files):
    st.header("Schmidt Surveys Analysis") # Changed from subheader
    schmidt_stop_words_orig = stop_words_nltk | {'comment','comments','could','would','one','think','also','get','us','like','see','provided','na','n/a','no','yes','please','emerson','schmidt','well','ovation','csa','field','service','material','replacement','system','systems','support','product','base','user','install','installed','the','and','is','in','to','of','a','for','with','on','that','it','as'}

    def schmidt_extract_tokens_orig(text): # Used for Inst Base Word Clouds originally
        if not isinstance(text, str): return []
        text = text.lower(); tokens = re.findall(r'\b[a-zA-Z]{2,}\b', text)
        minimal_stop_words = set(['comment','comments','could','would','one','think','the','and','is','in','to','of','a','for','with','on','that','it','as'])
        return [w for w in tokens if w not in minimal_stop_words]

    tab_titles = ["CSA Systems","Ovation Product Support","Field Service","Installed Base & Material Replacement"]
    tab1, tab2, tab3, tab4 = st.tabs(tab_titles)

    # --- Tab 1: CSA Systems (Original Snippet Logic, fixed Theme count) ---
    with tab1:
        st.header("CSA Systems Survey Insights") # Changed from subheader
        required_file = 'csa_systems_sentiment'
        if required_file not in files: st.warning(f"CSA Systems data ({required_file}) not found.")
        else:
            try:
                files[required_file].seek(0); df_csa = pd.read_excel(files[required_file]); df_csa = df_csa.copy()
                # --- Generate Bigrams ---
                negative_comments_csa = []; comment_cols_csa = [col for col in df_csa.columns if "O_Q" in str(col) and not str(col).lower().endswith(("_sentimentlabel", "_sentimentscore"))]
                for col in comment_cols_csa:
                    sentiment_col = col + "_SentimentLabel"
                    if sentiment_col in df_csa.columns:
                         try: neg_series = df_csa.loc[df_csa[sentiment_col] == "Negative", col].dropna().astype(str); negative_comments_csa.extend(neg_series.tolist())
                         except Exception: pass
                top_bigrams = []
                if negative_comments_csa:
                     tokens_csa = [];
                     for text in negative_comments_csa: words = [word.lower() for word in word_tokenize(text) if word.isalpha() and word.lower() not in schmidt_stop_words_orig and len(word) > 2]; tokens_csa.extend(words)
                     if len(tokens_csa) >= 2:
                          bigram_list_csa = list(ngrams(tokens_csa, 2))
                          if bigram_list_csa: top_bigrams = Counter(bigram_list_csa).most_common(40)
                # --- Plotting ---
                if not top_bigrams: st.info("No negative bigrams found for CSA visualizations.")
                else:
                    st.write("##### Negative Bigrams Word Cloud")
                    bigram_dict = {"_".join(b): f for b,f in top_bigrams} # Original dict creation
                    try:
                        wc = WordCloud(width=800, height=400, background_color="white", colormap="Reds").generate_from_frequencies(bigram_dict)
                        fig_wc, ax_wc = plt.subplots(figsize=(10,4)); ax_wc.imshow(wc, interpolation="bilinear"); ax_wc.axis("off"); ax_wc.set_title("Negative Bigrams Word-Cloud (CSA Systems Tech)", fontsize=14, fontweight="bold")
                        plt.tight_layout(); st.pyplot(fig_wc); plt.close(fig_wc)
                    except Exception as e: st.warning(f"Could not generate CSA Bigram Word Cloud: {e}")
            except Exception as e: st.error(f"Error during CSA Systems viz: {e}"); st.error(traceback.format_exc())

    # --- Tab 2: Ovation Product Support (Original Snippet Logic) ---
    with tab2:
        st.header("Ovation Product Support Survey Insights") # Changed from subheader
        required_file = 'ovation_support_sentiment'
        if required_file not in files: st.warning(f"Ovation Support data ({required_file}) not found.")
        else:
            # (Keep implementation as before)
            try:
                files[required_file].seek(0); df_ova = pd.read_excel(files[required_file]); df_ova = df_ova.copy()
                comment_fields_ova_orig = {'O_Q11': 'Q11: Dissatisfaction Reasons', 'O_Q20a': 'Q20a: Improvement Suggestions'}
                sentiment_suffix = "_SentimentLabel"; quarter_col_ova = 'Submit Quarter'
                def plot_wordcloud_ova(comments_series, title):
                    tokens = []; ova_stopwords = schmidt_stop_words_orig | {'comment','comments','provided','would','could'}
                    for txt in comments_series.dropna().astype(str).str.lower():
                         for w in word_tokenize(txt):
                              if w.isalpha() and w not in ova_stopwords: tokens.append(w)
                    if not tokens: st.info(f"Word Cloud '{title}': No words found."); return
                    freq = pd.Series(tokens).value_counts().to_dict()
                    try:
                         wc = WordCloud(width=800, height=400, background_color='white').generate_from_frequencies(freq)
                         fig, ax = plt.subplots(figsize=(10,5)); ax.imshow(wc, interpolation='bilinear'); ax.axis('off')
                         ax.set_title(title, fontweight='bold'); plt.tight_layout(); st.pyplot(fig); plt.close(fig)
                    except Exception as e: st.warning(f"Could not generate Ovation word cloud '{title}': {e}")
                st.write("##### Negative Feedback Word Clouds")
                for col, label in comment_fields_ova_orig.items():
                    if col in df_ova.columns and col + sentiment_suffix in df_ova.columns:
                         neg_texts = df_ova.loc[df_ova[col + sentiment_suffix] == 'Negative', col]
                         if not neg_texts.dropna().empty: plot_wordcloud_ova(neg_texts, f"Ovation Product Support Survey — Word Cloud: {label}")
                st.write("##### Negative Responses Over Time")
                time_trends = {}
                cols_for_trend_ova = ['O_Q11','O_Q20a','O_Q18a']
                if quarter_col_ova not in df_ova.columns: st.warning(f"Quarter column '{quarter_col_ova}' missing for Ovation trend.")
                else:
                     processed_trend = False
                     for col in cols_for_trend_ova:
                          sent_col = col + sentiment_suffix
                          if col in df_ova.columns and sent_col in df_ova.columns:
                               series = df_ova.loc[df_ova[sent_col]=='Negative'].groupby(quarter_col_ova).size()
                               name = "Q11: Dissatisfaction" if col=='O_Q11' else ("Q20a: Improvement" if col=='O_Q20a' else ("Q18a: Non-Renewal" if col=='O_Q18a' else col))
                               if not series.empty: time_trends[name] = series; processed_trend = True
                     if not processed_trend: st.info("No data for Ovation negative trend chart.")
                     else:
                          trend_df = pd.DataFrame(time_trends).fillna(0).astype(int)
                          try: trend_df = trend_df.sort_index(key=lambda x: x.str.extract(r'Q(\d)', expand=False).astype(float))
                          except: pass
                          fig_trend = plt.figure(figsize=(10,6)); ax_trend = fig_trend.add_subplot(111)
                          trend_df.plot(kind='bar', edgecolor='black', rot=45, legend=True, ax=ax_trend)
                          ax_trend.set_title("Ovation Product Support Survey —\nNegative Responses Over Time: Q11, Q20a & Q18a", fontweight='bold')
                          ax_trend.set_xlabel("Submit Quarter"); ax_trend.set_ylabel("Count of Negative Responses"); ax_trend.grid(axis='y', linestyle='--', alpha=0.5)
                          plt.tight_layout(); st.pyplot(fig_trend); plt.close(fig_trend)
            except Exception as e: st.error(f"Error during Ovation Support viz: {e}"); st.error(traceback.format_exc())

    # --- Tab 3: Field Service (Original Snippet Logic) ---
    with tab3:
        st.header("Field Service Survey Insights") # Changed from subheader
        required_file = 'field_service_sentiment'
        if required_file not in files: st.warning(f"Field Service data ({required_file}) not found.")
        else:
            # (Keep implementation as before)
            try:
                files[required_file].seek(0); df_fs = pd.read_excel(files[required_file]); df_fs = df_fs.copy()
                comment_fields_fs_orig = {'O_Q8': 'Q8: Improvement Suggestions', 'O_Q10a': 'Q10a: Value Justification', 'O_Q13a': 'Q13a: Likelihood-to-Use Reasons'}
                sentiment_suffix = "_SentimentLabel"; quarter_col_fs = 'Submit Quarter'
                def plot_wordcloud_fs(series, title):
                    tokens = []; fs_stopwords = schmidt_stop_words_orig | {'comment','comments','could','would','provided','also'}
                    for txt in series.dropna().astype(str).str.lower():
                         for w in word_tokenize(txt):
                              if w.isalpha() and w not in fs_stopwords: tokens.append(w)
                    if not tokens: st.info(f"Word Cloud '{title}': No words found."); return
                    freqs = pd.Series(tokens).value_counts().to_dict()
                    try:
                         wc = WordCloud(width=800, height=400, background_color='white').generate_from_frequencies(freqs)
                         fig, ax = plt.subplots(figsize=(10,5)); ax.imshow(wc, interpolation='bilinear'); ax.axis('off')
                         ax.set_title(f"Field Service Survey — Word Cloud: {title}", fontweight='bold'); plt.tight_layout(); st.pyplot(fig); plt.close(fig)
                    except Exception as e: st.warning(f"Could not generate FS word cloud '{title}': {e}")
                st.write("##### Negative Feedback Word Clouds")
                for col, friendly in comment_fields_fs_orig.items():
                     if col in ['O_Q8','O_Q10a'] and col in df_fs.columns and col + sentiment_suffix in df_fs.columns:
                          sel = df_fs.loc[df_fs[col + sentiment_suffix]=='Negative', col]
                          if not sel.dropna().empty: plot_wordcloud_fs(sel, friendly)
                st.write("##### Sentiment Distribution Across Key Questions")
                dist = []; processed_dist = False
                for col, friendly in comment_fields_fs_orig.items():
                     sent_col = col + sentiment_suffix
                     if sent_col in df_fs.columns:
                          pct = df_fs[sent_col].value_counts(normalize=True).reindex(['Positive','Neutral','Negative'], fill_value=0) * 100
                          if pct.sum() > 0: dist.append(pct.rename(friendly)); processed_dist = True
                if not processed_dist: st.info("No data for FS sentiment distribution chart.")
                else:
                     dist_df = pd.DataFrame(dist); colors = ['#4c78a8','#f58518','#e45756']
                     fig_dist, ax_dist = plt.subplots(figsize=(8,5))
                     dist_df.plot(kind='barh', stacked=True, color=colors, figsize=(8,5), edgecolor='black', ax=ax_dist, width=0.7)
                     ax_dist.set_title("Field Service Survey — Sentiment Distribution", fontweight='bold'); ax_dist.set_xlabel("Percentage of Responses")
                     ax_dist.legend(loc='lower right', title=None); ax_dist.invert_yaxis(); plt.tight_layout(); st.pyplot(fig_dist); plt.close(fig_dist)
                st.write("##### Negative Responses Over Time")
                time_trends = {}; processed_trend = False
                if quarter_col_fs not in df_fs.columns: st.warning(f"Quarter column '{quarter_col_fs}' missing for FS trend.")
                else:
                     for col, friendly in comment_fields_fs_orig.items():
                          sent_col = col + sentiment_suffix
                          if col in df_fs.columns and sent_col in df_fs.columns:
                              ser = df_fs.loc[df_fs[sent_col]=='Negative'].groupby(quarter_col_fs).size()
                              if not ser.empty: time_trends[friendly] = ser; processed_trend = True
                     if not processed_trend: st.info("No data for FS negative trend chart.")
                     else:
                          trend_df = pd.DataFrame(time_trends).fillna(0).astype(int)
                          try: trend_df = trend_df.sort_index(key=lambda x: x.str.extract(r'Q(\d)', expand=False).astype(float))
                          except: pass
                          fig_trend, ax_trend = plt.subplots(figsize=(10,6))
                          trend_df.plot(kind='bar', edgecolor='black', rot=45, ax=ax_trend, width=0.8)
                          ax_trend.set_title("Field Service Survey — Negative Responses Over Time", fontweight='bold'); ax_trend.set_xlabel("Submit Quarter"); ax_trend.set_ylabel("Count of Negative Responses")
                          ax_trend.legend(title=None); ax_trend.grid(axis='y', linestyle='--', alpha=0.5); plt.tight_layout(); st.pyplot(fig_trend); plt.close(fig_trend)
            except Exception as e: st.error(f"Error during Field Service viz: {e}"); st.error(traceback.format_exc())

    # --- Tab 4: Installed Base & Material Replacement (Original Snippet Logic - Inst Base Theme chart removed) ---
    with tab4:
        st.header("Installed Base & Material Replacement Insights") # Changed from subheader
        # --- Installed Base Visuals ---
        st.write("#### Installed User Base Survey") # Changed from header
        required_file_inst = 'installed_base_sentiment'
        if required_file_inst not in files: st.warning(f"Installed Base data ({required_file_inst}) not found.")
        else:
            try:
                files[required_file_inst].seek(0); df_inst = pd.read_excel(files[required_file_inst]); df_inst = df_inst.copy()
                # Re-define utilities from original snippet scope
                ib_stop_words = set(['comment','comments','could','would','one','think','the','and','is','in','to','of','a','for','with','on','that','it','as'])
                def ib_extract_tokens(text):
                     tokens = re.findall(r'\b[a-zA-Z]{2,}\b', str(text).lower()) # Ensure text is string
                     return [w for w in tokens if w not in ib_stop_words]
                # 1A) Satisfaction Distribution (Original Logic)
                st.write("##### Satisfaction Distribution")
                qs = ["Q2","Q4","Q6","Q9","Q11"]; labels = ["Installation Support","Startup Support","Follow-Up Support","Simulator Training","Training Satisfaction"]
                qs_present = [q for q in qs if q in df_inst.columns]
                if not qs_present: st.info("No relevant satisfaction questions found for Inst Base chart.")
                else:
                     sat_data = {}
                     for q, lbl in zip(qs, labels):
                          if q in qs_present:
                               df_inst[q] = pd.to_numeric(df_inst[q], errors='coerce')
                               mapped = df_inst[q].map(lambda x: 1 if pd.notna(x) and x>=4 else (0 if pd.notna(x) and x==3 else (-1 if pd.notna(x) and x<=2 else pd.NA))).dropna()
                               if not mapped.empty: pct = mapped.value_counts(normalize=True).reindex([1,0,-1]).fillna(0)*100; sat_data[lbl] = pct
                     if not sat_data: st.info("No valid satisfaction data for Inst Base chart.")
                     else:
                          sat = pd.DataFrame(sat_data).T; sat.columns = ["Positive %","Neutral %","Negative %"]
                          fig_sat, ax_sat = plt.subplots(figsize=(8,5))
                          sat.plot(kind='barh', stacked=True, color=["#4c78a8","#f58518","#e45756"], edgecolor='black', ax=ax_sat, width=0.7)
                          ax_sat.set_title("Installed Users: Satisfaction Distribution", pad=15); ax_sat.set_xlabel("Percentage (%)"); ax_sat.set_ylabel("")
                          ax_sat.tick_params(axis='y', which='major', pad=15, labelsize=9); plt.setp(ax_sat.get_yticklabels(), ha='right')
                          ax_sat.legend(loc='upper left', bbox_to_anchor=(1.02,1)); plt.subplots_adjust(right=0.75); plt.tight_layout(); st.pyplot(fig_sat); plt.close(fig_sat)
                # Word Cloud function specific to this section's needs (using original ib_extract_tokens)
                def plot_wordcloud_inst(text_series, title):
                    tokens = [t for txt in text_series.dropna().astype(str) for t in ib_extract_tokens(txt)] # Use IB specific token func
                    if not tokens: st.info(f"Word Cloud '{title}': No words."); return
                    freq = Counter(tokens)
                    try:
                        wc = WordCloud(width=800, height=400, background_color='white').generate_from_frequencies(freq) # Use original WC params
                        fig, ax = plt.subplots(figsize=(10,4)); ax.imshow(wc, interpolation='bilinear'); ax.axis('off'); ax.set_title(title); plt.tight_layout(); st.pyplot(fig); plt.close(fig)
                    except Exception as e: st.warning(f"Could not generate Inst Base word cloud '{title}': {e}")
                # 1B) Word Cloud – General Comments (Using original logic)
                neg_gen, gen_comment_col, gen_sent_col = pd.Series(dtype=str), "O_GEN_COM", "O_GEN_COM_SentimentLabel"
                if gen_comment_col in df_inst.columns and gen_sent_col in df_inst.columns: neg_gen = df_inst[df_inst[gen_sent_col]=="Negative"][gen_comment_col].dropna().astype(str)
                if not neg_gen.empty: st.write("##### Word Cloud: General Comments (Negative)"); plot_wordcloud_inst(neg_gen, "Word Cloud: General Comments")
                else: st.info("No negative General Comments for Word Cloud.")
                # 1C) Word Cloud – Additional Training Needs (Using original logic)
                neg_train, train_comment_col, train_sent_col = pd.Series(dtype=str), "O_Q11A", "O_Q11A_SentimentLabel"
                if train_comment_col in df_inst.columns and train_sent_col in df_inst.columns: neg_train = df_inst[df_inst[train_sent_col]=="Negative"][train_comment_col].dropna().astype(str)
                if not neg_train.empty: st.write("##### Word Cloud: Additional Training Needs (Negative)"); plot_wordcloud_inst(neg_train, "Word Cloud: Additional Training Needs")
                else: st.info("No negative Training Needs comments for Word Cloud.")
                # *** 1D) Themes chart REMOVED ***
            except Exception as e: st.error(f"Error during Installed Base viz: {e}"); st.error(traceback.format_exc())

        st.divider()
        # --- Material Replacement Visuals (Original Snippet Logic) ---
        st.write("#### Material Replacement Survey") # Changed from header
        required_file_mat = 'material_replacement_sentiment'
        if required_file_mat not in files: st.warning(f"Material Replacement data ({required_file_mat}) not found (check processing log).")
        else:
            try:
                files[required_file_mat].seek(0); df_mat = pd.read_excel(files[required_file_mat]); df_mat = df_mat.copy(); quarter_col_mat = 'Submit Quarter'
                # 2A) Satisfaction Distribution
                st.write("##### Satisfaction Distribution")
                mq = ["Q1","Q2","Q3","Q5","Q6","Q7","Q8"]; mlbl = ["Lead Time Quoted","Emergency Response","Delivery Commitment","Part Availability","Personnel Attitude","Ease of Doing Business","Problem Resolution"]
                mq_present = [q for q in mq if q in df_mat.columns]
                if not mq_present: st.info("No relevant satisfaction questions found for Mat Repl chart.")
                else:
                     msat_data = {}
                     for q, lbl in zip(mq, mlbl):
                          if q in mq_present:
                               df_mat[q] = pd.to_numeric(df_mat[q], errors='coerce')
                               mapped = df_mat[q].map(lambda x: 1 if x>=4 else (0 if x==3 else -1 if x is not None else pd.NA)).dropna()
                               if not mapped.empty: pct = mapped.value_counts(normalize=True).reindex([1,0,-1]).fillna(0)*100; msat_data[lbl] = pct
                     if not msat_data: st.info("No valid satisfaction data for Mat Repl chart.")
                     else:
                          msat = pd.DataFrame(msat_data).T; msat.columns = ["Positive %","Neutral %","Negative %"]
                          fig_msat, ax_msat = plt.subplots(figsize=(8,5))
                          msat.plot(kind='barh', stacked=True, color=["#4c78a8","#f58518","#e45756"], edgecolor='black', ax=ax_msat, width=0.7)
                          ax_msat.set_title("Material Replacement: Satisfaction Distribution", pad=15); ax_msat.set_xlabel("Percentage"); ax_msat.set_ylabel("")
                          ax_msat.tick_params(axis='y', which='major', pad=15, labelsize=9); plt.setp(ax_msat.get_yticklabels(), ha='right')
                          ax_msat.legend(loc='upper left', bbox_to_anchor=(1.02,1)); plt.subplots_adjust(right=0.75); plt.tight_layout(); st.pyplot(fig_msat); plt.close(fig_msat)
                # 2B) Negative Responses Over Time
                st.write("##### Negative Responses Over Time")
                mat_trend_map_orig = {"Service Impr.": "O_Q10", "Value Concerns": "O_Q12a", "Recommend Hesitance": "O_Q16a"}
                time_trends_mat = {}; processed_trend = False
                if quarter_col_mat not in df_mat.columns: st.warning(f"Quarter column '{quarter_col_mat}' missing for Mat Repl trend.")
                else:
                     for label, col_key in mat_trend_map_orig.items():
                          sentiment_col = col_key + "_SentimentLabel"
                          if col_key in df_mat.columns and sentiment_col in df_mat.columns:
                               series = df_mat.loc[df_mat[sentiment_col]=="Negative"].groupby(quarter_col_mat).size()
                               if not series.empty: time_trends_mat[label] = series; processed_trend = True
                     if not processed_trend: st.info("No data for Mat Repl negative trends.")
                     else:
                          trend_df = pd.DataFrame(time_trends_mat).fillna(0).astype(int)
                          try: trend_df = trend_df.sort_index(key=lambda x: x.str.extract(r'Q(\d)', expand=False).astype(float))
                          except: pass
                          fig_trend, ax_trend = plt.subplots(figsize=(8,4)) # Use original figsize
                          trend_df.plot(kind='bar', figsize=(8,4), edgecolor='black', ax=ax_trend, width=0.8) # Use original plot call
                          ax_trend.set_title("Negative Responses Over Time: Material Replacement"); ax_trend.set_xlabel("Submit Quarter"); ax_trend.set_ylabel("Count")
                          plt.xticks(rotation=45, ha='right') # Original rotation
                          plt.tight_layout(); st.pyplot(fig_trend); plt.close(fig_trend)
            except Exception as e: st.error(f"Error during Material Repl viz: {e}"); st.error(traceback.format_exc())

# --- Streamlit App Layout ---

# Initialize session state variables if they don't exist
if 'processing_started' not in st.session_state:
    st.session_state['processing_started'] = False
if 'processed_data' not in st.session_state:
    st.session_state['processed_data'] = {}
if 'processing_done' not in st.session_state:
    st.session_state['processing_done'] = False

# --- Logo ---
# Replace with path or URL to your logo
logo_path = "emerson_logo.png"
# Use columns to control width/placement if needed
col1_logo, col2_logo = st.columns([1, 5]) # Adjust ratio as needed
with col1_logo:
    try:
        st.image(logo_path, width=250) # Adjust width as needed
    except Exception as img_e:
        st.warning(f"Could not load logo: {img_e}. Please replace placeholder path.")

with col2_logo:
    st.markdown("<h1 style='text-align: left-center;'>📊 Customer Survey Analysis Dashboard</h1>", unsafe_allow_html=True)

# --- Introduction / Instructions ---
if not st.session_state.processing_started:
    st.markdown("---")
    st.markdown("## About This Dashboard") # Use ## or ### for desired size
    with st.expander("Click to see details", expanded=True):
        st.markdown("""
        ### *This dashboard consolidates raw survey data to highlight customer sentiment, key themes, and performance trends across services and interactions.*
                    
        ### <u> **To get started:** </u>
        ### 1.  Use the sidebar to upload all required raw survey Excel files.
        ### 2.  Verify that all files are recognized correctly.
        ### 3.  Click the 'Start Processing and Analysis' button in the sidebar.
        """, unsafe_allow_html=True)
    st.info("Please upload files and start processing using the sidebar.")
    st.markdown("---")


# --- Sidebar ---
st.sidebar.header("1. Upload Raw Data Files")
expected_files_config = { "FAT": {"name": "FAT.xlsx", "is_quarterly": False}, "AEIC": {"name": "AEIC.xlsx", "is_quarterly": False}, "CSA_Q": {"pattern": r"System Q\d Final Data File\.xlsx", "is_quarterly": True, "base_name": "CSA Systems"}, "Ovation_Q": {"pattern": r"Schmidt Ovation Product Support Q\d Final Data\.xlsx", "is_quarterly": True, "base_name": "Ovation Support"}, "Material_Q": {"pattern": r"Schmidt Material Replacement Q\d Final Data File\.xlsx", "is_quarterly": True, "base_name": "Material Replacement"}, "FieldService_Q": {"pattern": r"Schmidt Field Service Q\d Final Data File\.xlsx", "is_quarterly": True, "base_name": "Field Service"}, "InstalledBase": {"name": "Schmidt 2024 Installed Base Final Data File.xlsx", "is_quarterly": False}, "NAUsersWritten": {"name": "NAUsersWrittenParsed.xlsx", "is_quarterly": False}, "NAUsersTemplate": {"name": "NA_Users_Group_Template_Final.xlsx", "is_quarterly": False} }
uploaded_files = st.sidebar.file_uploader("Select all required Excel files (.xlsx)", type="xlsx", accept_multiple_files=True, help="Upload all raw survey files.")

# --- Main Processing Logic ---
if uploaded_files:
    st.sidebar.success(f"{len(uploaded_files)} file(s) uploaded.")
    upload_status = {}; uploaded_files_map = {}
    for file in uploaded_files:
        matched = False
        for key, config in expected_files_config.items():
            if config["is_quarterly"]:
                if re.match(config["pattern"], file.name, re.IGNORECASE):
                    if key not in uploaded_files_map: uploaded_files_map[key] = []
                    uploaded_files_map[key].append(file); upload_status[key] = upload_status.get(key, 0) + 1; matched = True
            elif "name" in config and file.name.lower() == config["name"].lower():
                 uploaded_files_map[key] = file; upload_status[key] = 1; matched = True; break
    missing_files_errors = []
    for key, config in expected_files_config.items():
         if key not in upload_status: missing_files_errors.append(f"- Missing: {config.get('name', config.get('pattern'))} (Key: {key})")
         elif config["is_quarterly"] and upload_status.get(key, 0) < 4: missing_files_errors.append(f"- Incomplete: Expected 4 for {key}, found {upload_status.get(key, 0)}.")
    if missing_files_errors: st.sidebar.error("Upload Errors:\n" + "\n".join(missing_files_errors));
    else: st.sidebar.success("All expected files/patterns uploaded.")

    if st.sidebar.button("🚀 Start Processing and Analysis"):
        st.session_state['processing_started'] = True # Set flag to hide initial instructions
        st.session_state['processed_data'] = {} # Clear previous results
        st.session_state['processing_done'] = False

        num_cleaning = 8; num_sentiment = 8; num_summary = 8
        total_steps = num_cleaning + num_sentiment + num_summary
        progress_bar = st.progress(0, text="Starting Process...")
        current_step = 0

        try:
            processed_data_buffers = {}
            st.sidebar.info("🔄 Step 1/4: Cleaning data...") # Show status on main page briefly
            key_to_base_name = {"FAT": "fat","AEIC": "aeic","CSA_Q": "csa_systems","Ovation_Q": "ovation_support","Material_Q": "material_replacement","FieldService_Q": "field_service","InstalledBase": "installed_base","NAUsersWritten": "na_users_written"}
            cleaning_tasks = [("FAT", clean_fat_data, False),("AEIC", clean_aeic_data, False),("CSA_Q", clean_csa_systems_data, True),("Ovation_Q", clean_ovation_product_support_data, True),("Material_Q", clean_material_replacement_data, True),("FieldService_Q", clean_field_service_data, True),("InstalledBase", clean_installed_base_data, False),("NAUsersWritten", clean_na_users_written_data, False)]
            for key, clean_func, is_quarterly in cleaning_tasks:
                 base_name = key_to_base_name[key]; output_key = f"{base_name}_cleaned"
                 task_failed = False; msg = f"Cleaning {base_name}"; df_cleaned = None
                 if key not in uploaded_files_map: task_failed = True; msg += " (Not Uploaded)"
                 else:
                      input_obj = uploaded_files_map[key]
                      try:
                          if key == "NAUsersWritten": df_cleaned = clean_func(input_obj)
                          elif is_quarterly:
                               df_combined = read_combine_quarterly_files(input_obj, base_name)
                               if df_combined is not None: df_cleaned = clean_func(df_combined)
                               else: task_failed = True; msg += " (Combine Fail)"
                          else:
                              df_raw = read_excel_from_upload(input_obj)
                              if df_raw is not None: df_cleaned = clean_func(df_raw)
                              else: task_failed = True; msg += " (Read Fail)"
                          if not task_failed and df_cleaned is not None:
                               buffer = save_excel_to_buffer(df_cleaned)
                               if buffer: processed_data_buffers[output_key] = buffer
                               else: task_failed = True; msg += " (Save Fail)"
                          elif not task_failed: task_failed = True; msg += " (Clean Fail)"
                      except Exception as clean_e: task_failed=True; msg+=f" (Error: {clean_e})"
                 current_step += 1; progress_value = min(1.0, current_step / total_steps)
                 progress_bar.progress(progress_value, text=msg)
                 if task_failed: st.sidebar.warning(f"⚠️ {msg}") # Use sidebar for persistent errors

            if "NAUsersTemplate" in uploaded_files_map:
                try: template_file = uploaded_files_map["NAUsersTemplate"]; template_file.seek(0); processed_data_buffers["na_users_template_raw"] = io.BytesIO(template_file.getvalue())
                except Exception as e: st.sidebar.error(f"❌ Failed load NA Template: {e}")
            else: st.sidebar.warning("⚠️ NA Users Template missing.")

            cleaned_files_keys = [k for k in processed_data_buffers if k.endswith('_cleaned')]
            if not cleaned_files_keys: st.error("No data successfully cleaned. Stopping."); st.stop()

            st.sidebar.info("🔄 Step 2/4: Applying sentiment analysis...")
            sentiment_keys = []
            for key in cleaned_files_keys:
                 base_name = key.replace('_cleaned', ''); task_failed = False; msg = f"Sentiment: {base_name}"
                 if key in processed_data_buffers:
                     buffer = processed_data_buffers[key]; buffer.seek(0)
                     try:
                         df_cleaned = pd.read_excel(buffer); df_sentiment = apply_sentiment_analysis(df_cleaned)
                         if df_sentiment is not None:
                             sentiment_key = base_name + '_sentiment'
                             sent_buffer = save_excel_to_buffer(df_sentiment)
                             if sent_buffer: processed_data_buffers[sentiment_key] = sent_buffer; sentiment_keys.append(sentiment_key)
                             else: task_failed = True; msg += " (Save Fail)"
                         else: task_failed = True; msg += " (Analysis Fail)"
                     except Exception as e: task_failed=True; msg+=f" (Error: {e})"
                 else: task_failed=True; msg+="(Clean step failed?)"
                 current_step += 1; progress_value = min(1.0, current_step / total_steps)
                 progress_bar.progress(progress_value, text=msg)
                 if task_failed: st.sidebar.warning(f"⚠️ {msg}")

            st.sidebar.info("🔄 Step 3/4: Generating sentiment summaries...")
            summary_keys = []
            for key in sentiment_keys: # Only process successful sentiment files
                 base_name = key.replace('_sentiment', ''); task_failed = False; msg = f"Summarizing: {base_name}"
                 if key in processed_data_buffers:
                     buffer = processed_data_buffers[key]; buffer.seek(0)
                     try:
                         df_sentiment = pd.read_excel(buffer)
                         group_col = "Submit Quarter";
                         if 'Submit_Quarter_Num' in df_sentiment.columns: group_col = 'Submit_Quarter_Num'
                         elif 'Submit Quarter_Num' in df_sentiment.columns: group_col = 'Submit Quarter_Num'
                         df_summary = generate_sentiment_summary(df_sentiment, groupby_col=group_col)
                         if df_summary is not None and not df_summary.empty:
                             summary_key = base_name + '_summary'
                             summary_buffer = save_excel_to_buffer(df_summary)
                             if summary_buffer:
                                 formatted_summary_buffer = adjust_excel_column_widths(summary_buffer)
                                 processed_data_buffers[summary_key] = formatted_summary_buffer; summary_keys.append(summary_key)
                             else: task_failed = True; msg += " (Save Fail)"
                         else: task_failed = True; msg += " (Gen Fail/Empty)"
                     except Exception as e: task_failed=True; msg+=f" (Error: {e})"
                 else: task_failed=True; msg+="(Prev Step Failed?)"
                 current_step += 1; progress_value = min(1.0, current_step / total_steps)
                 progress_bar.progress(progress_value, text=msg)
                 if task_failed: st.sidebar.warning(f"⚠️ {msg}")

            st.sidebar.info("🔄 Step 4/4: Preparing visualizations...")
            st.session_state['processed_data'] = processed_data_buffers
            st.session_state['processing_done'] = True
            progress_bar.progress(1.0, text="Processing Complete!")
            st.success("✅ Processing Complete! Results shown below.")

        except Exception as main_e:
             st.error(f"Critical Error during processing: {main_e}"); st.error(traceback.format_exc())
             st.session_state['processing_done'] = False


# --- Display Results ---
if st.session_state.get('processing_done', False):
    st.markdown("---"); st.header("📊 Visualization Results")
    processed_data = st.session_state.get('processed_data', {})

    # Use radio buttons to select which survey viz to show
    survey_options = ["FAT", "AEIC", "NA Users", "Schmidt", "Hide All"]
    selected_survey = st.radio(
        "Select Survey Visualization to Display:",
        options=survey_options,
        index=len(survey_options)-1, # Default to "Hide All"
        horizontal=True,
        label_visibility="collapsed" # Hide the label "Select Survey..." if desired
    )

    st.markdown("---") # Separator

    # Conditionally display the selected survey visualization
    if selected_survey == "FAT":
        if 'fat_summary' in processed_data and 'fat_sentiment' in processed_data:
            display_fat_visualizations(processed_data['fat_summary'], processed_data['fat_sentiment'])
        else:
            st.warning("FAT results unavailable (check processing warnings).")

    elif selected_survey == "AEIC":
        if 'aeic_sentiment' in processed_data:
            display_aeic_visualizations(processed_data['aeic_sentiment'])
        else:
            st.warning("AEIC results unavailable (check processing warnings).")

    elif selected_survey == "NA Users":
        if 'na_users_written_sentiment' in processed_data and 'na_users_template_raw' in processed_data:
            display_na_users_visualizations(processed_data['na_users_written_sentiment'], processed_data['na_users_template_raw'])
        else:
            st.warning("NA Users results unavailable (check processing warnings).")

    elif selected_survey == "Schmidt":
        # Check for Schmidt components logic (remains the same)
        schmidt_base_keys_map = {"csa_systems": "csa_systems_sentiment","ovation_support": "ovation_support_sentiment","field_service": "field_service_sentiment","installed_base": "installed_base_sentiment","material_replacement": "material_replacement_sentiment"}
        schmidt_files_for_viz = {}
        missing_schmidt = []
        all_schmidt_found = True
        for base_key, expected_sentiment_key in schmidt_base_keys_map.items():
            if expected_sentiment_key not in processed_data:
                missing_schmidt.append(base_key.replace("_", " ").title())
                all_schmidt_found = False
            else:
                schmidt_files_for_viz[expected_sentiment_key] = processed_data[expected_sentiment_key]

        if all_schmidt_found:
            # The display function creates the tabs *within* this selected section
            display_schmidt_visualizations_original(schmidt_files_for_viz)
        else:
            st.error(f"Cannot display full Schmidt visualizations. Missing processed data for: {', '.join(missing_schmidt)}")
            st.info("Please check the warnings generated during processing for errors related to these surveys.")

    # elif selected_survey == "Hide All":
        # Do nothing, show no charts

# --- Handle states before processing is done ---
# (This part remains unchanged)
elif not uploaded_files and not st.session_state.get('processing_started', False):
    pass
elif uploaded_files and not st.session_state.get('processing_started', False):
     st.info("Please click 'Start Processing and Analysis' in the sidebar.")


# --- Handle states before processing is done ---
elif not uploaded_files and not st.session_state.get('processing_started', False):
    # Initial instructions shown only if files haven't been uploaded AND processing hasn't started
    # Instructions are now shown at the top by default if not processing_started
    pass
elif uploaded_files and not st.session_state.get('processing_started', False):
    # Case where files are uploaded, but processing hasn't been started yet
    st.info("Please click 'Start Processing and Analysis' in the sidebar.")

# --- Footer/Sidebar Instructions (Displayed *after* processing starts) ---
# (This block remains unchanged at the very end of the file)
if st.session_state.get('processing_started', False):
    st.sidebar.markdown("---")
    st.sidebar.markdown("### Instructions Summary")
    st.sidebar.markdown("""
    #### 1. **Upload Files:** Use the 'Browse files' button above.
    #### 2. **Verify:** Check messages ensure files were recognized.
    #### 3. **Process Data:** Click 'Start Processing'.
    #### 4. **View Results:** Charts appear on main page. Check warnings if issues occur.
    """)
# --- End of Script ---